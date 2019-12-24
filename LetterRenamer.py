# This application uses the Letter to Development Workbook (Provided by Financial Assistance) to rename all
# provided pdf file within a selected folder to the appropriate format suited for the Foundation
# Thanh Tran - June 10, 2019

from openpyxl import load_workbook
import os
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime

startProgram = False
filesNotRenamed = []
listOfChar = ['/', '\\', '?', '%', '*', ':', '|', '"', '<', '>']
# check for quotes and add \ to allow appending to list
def convertQuotes(word):
    if word is not None and isinstance(word, str):
        if '"' in word:
            return word.replace('"', '\"')
        else:
            return word
    else:
        return word
# convert fund name to an accepted file name for windows
def fundNameConverter(word):
    for char in listOfChar:
        if char in word:
            return word.replace(char, '-')
    return word

def start_button():
    global startProgram
    startProgram = True
    root.quit()

# button to browse and choose file
def filebrowse_button():
    file = filedialog.askopenfile(parent=root, mode='rb', title='Select a file', filetypes=[('Excel file', '*.xlsx;*.xlsm;*xls')])
    if file is not None:
        letterToDevFile.set(file.name.replace("/", "\\"))

# button to browse and choose folder
def browse_button():
    path = filedialog.askdirectory(title='Select a folder')
    if path is not None:
        letterFolder.set(path.replace("/", "\\"))

def about_window():
    # initialize about window
    windows = Toplevel(root)
    windows.title('Foundation Letter Renamer')
    topframe = Frame(windows)
    topframe.pack(side=TOP)
    bottomframe = Frame(windows)
    bottomframe.pack(side=TOP)
    closeframe = Frame(windows)
    closeframe.pack(side=TOP)
    Label(topframe, text='About', font=('Aria', 12, 'bold')).grid(row=0, padx=10, pady=20, columnspan=3)
    Label(topframe, text=
        'This application uses the General Select Workbook (Provided by Financial Assistance) '
        'to rename all provided pdf file within a selected folder to the appropriate format suited for the Foundation. When "Start" is pressed'
        ' please allow some time for the program to process as window will be in a "Not Responding State". '
        'You will be alerted when the renaming have been completed'
        , wraplength=550, justify='left').grid(row=1, padx=25,pady=10)
    Label(topframe, text='Information regarding initial configuration:\n', font=('Aria', 9, 'bold')).grid(row=2, column=0, sticky='w', padx=25)
    Label(bottomframe, text='●').grid(padx=(0, 5), row=0, column=0)
    Label(bottomframe, text=
        'Student Number Column, Fund ID Column and Fund Title Column are set at their default values, only change if the positioning '
        'of the column have changed in the workbook. For example, Fund ID is found in column "A" which is the first column, '
        'hence 1 being in the "Fund ID Column field"', wraplength=500, justify='left').grid(row=0, column=1, sticky='w')
    Label(bottomframe, text='●').grid(padx=(0, 5), row=1, column=0)
    Label(bottomframe, text='Letter to Development: browse and locate Letter to Development File', wraplength=500, justify='left').grid(row=1, column=1, sticky='w')
    Label(bottomframe, text='●').grid(padx=(0, 5), row=2, column=0)
    Label(bottomframe, text='Folder with Letters: browse and select folder containing all letters, subfolders are selected in the process', wraplength=500, justify='left').grid(row=2, column=1, sticky='w')
    Label(closeframe, text=
        'Files that have not been renamed, will be logged in a .log file found in the folder "Files Unchanged Logs".'
        ' The log will provide direct folder path and file, so the user can easily locate and rename the file.'
          , wraplength=550, justify='left').grid(row=0, padx=25,pady=10)
    Button(closeframe, text='Close', command=windows.destroy, width=15).grid(row=1, column=0, pady=25)

# used to select calculated file to
root = Tk()
root.title('Foundation Letter Renamer')
root.tk.call('wm', 'iconphoto', root._w, PhotoImage(file='mohawklogo.gif'))
# label section
Label(root, text='Letters to Development', font=('Aria', 12, 'bold')).grid(row=0, padx=10, pady=20, columnspan=3)
Label(root, text='Student Number Column: ').grid(row=1, padx=10, pady=5, sticky='e')
Label(root, text='First Name Column:').grid(row=2, padx=10, pady=10, sticky='e')
Label(root, text='Last Name Column: ').grid(row=3, padx=10, pady=10, sticky='e')
Label(root, text='Fund Title Column:').grid(row=4, padx=10, pady=10, sticky='e')
Label(root, text='Workbook: ').grid(row=5, padx=10, pady=10, sticky='e')
Label(root, text='Folder with Letters: ').grid(row=6, padx=10, pady=10, sticky='e')
# default value section
studentNumColValue = StringVar()
studentNumColValue.set('Student #')
firstNameColValue = StringVar()
firstNameColValue.set('First Name')
lastNameColValue = StringVar()
lastNameColValue.set('Last Name')
fundTitleColValue = StringVar()
fundTitleColValue.set('Award Name')
letterToDevFile = StringVar()
letterFolder = StringVar()
# input section
Entry(root, textvariable=studentNumColValue, width=45).grid(row=1, column=1)
Entry(root, textvariable=firstNameColValue, width=45).grid(row=2, column=1)
Entry(root, textvariable=lastNameColValue, width=45).grid(row=3, column=1)
Entry(root, textvariable=fundTitleColValue, width=45).grid(row=4, column=1)
Entry(root, textvariable=letterToDevFile, width=45, state="readonly").grid(row=5, column=1)
Entry(root, textvariable=letterFolder, width=45, state="readonly").grid(row=6, column=1)
# button section
Button(root, text='Browse', command=filebrowse_button, width=15).grid(row=5, column=2, pady=10, padx=10)
Button(root, text='Browse', command=browse_button, width=15).grid(row=6, column=2, pady=10, padx=10)
Button(root, text='Start', command=start_button, width=15, bg="medium sea green", fg="white").grid(row=8, column=1, pady=10, padx=10)
Button(root, text='Help', command=about_window, width=15).grid(row=9, column=2, pady=10, padx=10)
root.mainloop()

# Main program
if letterToDevFile.get() != "" and letterFolder.get() != "" and startProgram is True:
    academicYear = ""
    StudentInfo = {}
    i = 0
    headerCompleted = False
    wb_obj = load_workbook(filename=letterToDevFile.get(), data_only=True, read_only=True)
    sheetsName = wb_obj.sheetnames
    for sheets in sheetsName:
        if sheets.find('GEN.SEL') >= 0:
            wsheet =wb_obj[sheets]
            academicYearSplit = sheets.split(' ')
            academicYear = academicYearSplit[0]
            print(academicYear)
            for values in wsheet.iter_rows(min_row=1):
                results = []
                for v in values:
                    results.append(convertQuotes(v.value))
                if headerCompleted is True:
                    if ( results[int(list(StudentInfo.values())[0].index(studentNumColValue.get()))] != 0 ):
                        StudentInfo[results[int(list(StudentInfo.values())[0].index(studentNumColValue.get()))]] = results
                    
                else:
                    StudentInfo[0] = results
                    print(StudentInfo[0])
                headerCompleted = True
   
    # keep track of file change
    counter = 0

    try:
        firstNameCol = int(list(StudentInfo.values())[0].index(firstNameColValue.get()))
        lastNameCol = int(list(StudentInfo.values())[0].index(lastNameColValue.get()))
        fundTitleCol = int(list(StudentInfo.values())[0].index(fundTitleColValue.get()))
        studentNumCol = int(list(StudentInfo.values())[0].index(studentNumColValue.get()))
    except ValueError:
        messagebox.showerror("Error", "Heading was not found in latest two worksheet, please align and adjust")
    # cycle through all files within specified directory
    for subdir, dirs, files in os.walk(letterFolder.get()):
        for file in files:
            try:
                studentNumber = file.split("- ")
            except ValueError:
                continue
            else:
                print(studentNumber[len(studentNumber) - 1])
                if studentNumber[len(studentNumber) - 1].endswith('pdf'):
                    try:
                        importedName = studentNumber[1].strip()
                        
                        studentNumber = int(importedName[-9:])
                        print(studentNumber)
                        if studentNumber in StudentInfo.keys():
                            os.rename(os.path.join(subdir, file), os.path.join(subdir, academicYear + " "
                                + fundNameConverter(str(StudentInfo.get(studentNumber)[fundTitleCol]))) + " ( " + str(StudentInfo.get(studentNumber)[lastNameCol]) + "." + str(StudentInfo.get(studentNumber)[firstNameCol]) + " )" + ".pdf")
                            counter += 1
                        else:
                            filesNotRenamed.append(os.path.join(subdir, file))
                    except Exception as ex:
                        # list file location that haven't been changed
                        filesNotRenamed.append(os.path.join(subdir, file))
                        continue
                else:
                    continue
    # log files that haven't been written
    if len(filesNotRenamed) > 0:
        with open('Files Unchanged Logs\\files_not_changed' + datetime.today().strftime('%Y%m%d-%H%M') + '.log', 'w') as f:
            for item in filesNotRenamed:
                f.write("%s\n" % item)
    messagebox.showinfo('Success', str(counter) + ' files have been changed and rename')
elif letterToDevFile.get() == "" or letterFolder.get() == "":
    messagebox.showinfo('Error', 'File and folder not select, Please run program again')
